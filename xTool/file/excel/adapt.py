from collections import defaultdict
from io import BytesIO
from typing import Any, Dict, List, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

__all__ = [
    "adapt_sheet_weight_height",
    "to_list",
    "to_matrix",
    "from_dict",
]


def adapt_sheet_weight_height(sheet: Worksheet, first_header_row: int = 1):
    """
    自动调整 Excel 工作表的行高和列宽，以确保内容能够完整显示。使单元格内容（包括换行符 \n 分割的多行文本）能够完整显示。

    :param sheet: 要调整的 Excel 工作表对象（来自 openpyxl）
    :param first_header_row (int, 默认 1)：表头开始的行数（前几行可能是注释或其他元数据，不参与列宽计算）
    """
    # 列宽计算公式: (max_str_len).encode("gbk") * 1.3 (同一单元格中通过\n分割的字符视为独立的长度)
    # 注：这里为什么用gbk编码呢？因为在gbk中一个中文字符为两个字节，英文为1个字节，而在excel中一个中文字符的大小接近于两个英文小写字符
    # 因此用gbk编码来统计字符大小作为字符的宽度。
    # 事实上大写英文字符，小数点，特殊符号大小也各不相同，但是openpyxl中没有自适应宽度的函数

    # 计算最大列数和初始化列宽数组
    row_num, col_num = sheet.max_row, sheet.max_column
    # 初始化一个长度为 col_num + 1 的列表（索引从 1 开始），用于存储每列的最大宽度。
    max_col_dimensions: List[int] = [0] * (col_num + 1)  # 初始化列宽数组

    # 遍历所有行和列
    for row in range(first_header_row, row_num + 1):
        for col in range(1, col_num + 1):
            # 设置单元格自动换行（1-based）
            cell = sheet.cell(row, col)
            # 设置单元格对齐方式为自动换行（多行文本时自动调整行高）。
            cell.alignment = Alignment(wrapText=True)

            # 计算单元格字符宽度
            try:
                cell_str_len_list = [len(cell_str.encode("gbk")) for cell_str in str(cell.value or "").split("\n")]
                max_col_dimensions[col] = max(max_col_dimensions[col], max(cell_str_len_list) * 1.3)
            except UnicodeEncodeError:
                # 回退到 UTF-8（不精确，但避免崩溃）
                cell_str_len_list = [len(str(cell.value or "")) * 1.5]  # 粗略估计
                max_col_dimensions[col] = max(max_col_dimensions[col], cell_str_len_list[0])

    # 遍历所有列并设置宽度
    for col in range(1, col_num + 1):
        sheet.column_dimensions[get_column_letter(col)].width = max_col_dimensions[col]


def to_list(excel: BytesIO, header_row: int = 0, sheet_name: str = "") -> List[Dict]:
    """解析excel文件为数据字典

    :param excel: 表示 Excel 文件的二进制数据流。BytesIO 是 Python 中用于处理二进制数据的类，通常用于内存中的文件操作。
    :param header_row(int, 默认 0)：指定表头所在的行数。注意，Python 使用零基索引，因此 header_row=0 表示第一行是表头。
    :param sheet_name (str, 默认 "")：指定要解析的工作表名称。如果为空字符串，则默认解析活动工作表（即第一个工作表）。
    """
    # 打开文件
    try:
        if not sheet_name:
            workbook = load_workbook(excel)
            sheet = workbook.active
        else:
            workbook = load_workbook(excel)
            sheet = workbook[sheet_name]
    except KeyError:
        raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
    excel_rows = list(sheet.rows)

    # 提取表头
    if header_row >= len(excel_rows):
        raise ValueError(f"Header row index {header_row} is out of range.")
    header_list = []
    for header in excel_rows[header_row]:
        value = header.value
        # 可选：处理空表头，例如使用默认值或跳过
        header_list.append(value if value is not None else f"Column_{header_list.__len__() + 1}")

    # 遍历数据行并构建字典
    excel_data_dict_list = []
    for content_row in excel_rows[header_row + 1 :]:
        # 跳过空行
        if all(content.value is None for content in content_row):
            continue

        # 获得一行中每个单元格的值
        content_list = []
        for cell in content_row:
            value = cell.value
            # 处理 None 值
            content_list.append(str(value) if value is not None else "")

        # 检查列数是否匹配
        if len(content_list) != len(header_list):
            raise ValueError(f"Row {content_row[0].row} has {len(content_list)} columns, expected {len(header_list)}.")

        excel_data_dict_list.append(dict(zip(header_list, content_list)))

    return excel_data_dict_list


def to_matrix(excel: Union[BytesIO, str]) -> Dict:
    """
    解析Excel文件为二维矩阵，默认第一行和第一列是索引指标。

    :param excel: Excel二进制文件（BytesIO）或文件路径（str）
    :return: 嵌套字典，结构为 {列索引: {行索引: 值}}
    :raises ValueError: 若文件不存在或格式无效
    """
    try:
        wb = load_workbook(excel) if isinstance(excel, BytesIO) else load_workbook(filename=excel)
    except FileNotFoundError:
        raise ValueError("Excel file not found.")
    except InvalidFileException:
        raise ValueError("Invalid Excel file format.")
    except Exception as e:
        raise ValueError(f"Failed to load Excel file: {str(e)}")

    matrix_data = {}

    # 遍历所有工作表，sheet是1-based
    for sheet in wb.worksheets:
        # 获取列索引（第一行，跳过A1）
        # sheet[1] 获取第一行。
        # [1:] 切片操作，跳过第一个单元格（即 A1），从 B1 开始。
        # 使用列表推导式遍历第一行的每个单元格，提取其 value 属性，如果 value 不为 None，则包含在列表中
        columns = [cell.value for cell in sheet[1][1:] if cell.value is not None]

        # 获取行索引（第一列，跳过A1）
        rows = [
            sheet.cell(row=row_idx, column=1).value
            for row_idx in range(2, sheet.max_row + 1)
            if sheet.cell(row=row_idx, column=1).value is not None
        ]

        # 确保行列索引存在
        if not columns or not rows:
            continue  # 跳过没有有效行列索引的工作表

        # 遍历数据行和列
        for sheet_idx, sheet in enumerate(wb.worksheets):
            matrix_data[sheet_idx] = defaultdict(dict)
            for row in range(2, sheet.max_row + 1):
                for col in range(2, sheet.max_column + 1):
                    if sheet.cell(row, col).value is None:
                        continue
                    col_key = sheet.cell(1, col).value
                    row_key = sheet.cell(row, 1).value
                    matrix_data[sheet_idx][col_key][row_key] = sheet.cell(row, col).value

    matrix_data = {key: dict(value) for key, value in matrix_data.items()}

    return matrix_data


def from_dict(
    data_dict__list: List[Dict],
    template: str = None,
    headers: List[Union[str, Dict[str, Any]]] = None,
    header_style: Dict[str, str] = None,
    match_header: bool = False,
) -> Workbook:
    """
    将数据字典序列化为excel对象

    :param data_dict__list: 数据字典列表，每个字典代表一行数据
    :param template: excel模板路径（优先以模板的头部样式作为excel的头部）
    :param headers: excel数据头，格式为 [{"id": "header_id", "name": "header_name"}] 或 ["header1", "header2"]
    :param header_style: excel的头部样式（颜色），格式为 {"header_id": "color_code"}
    :param match_header: 数据是否匹配表头，如果为True，则根据 header 严格匹配列名，若不存在，则在该 cell 填充空
    :return: 返回一个Workbook对象
    """
    wb: Workbook = Workbook()
    sheet: Worksheet = wb.active
    # 数据起始行（第1行为表头）
    first_data_row: int = 2

    if template:
        try:
            wb = load_workbook(template)
            sheet = wb.active
            first_data_row = 3  # 模板第1行为注释，第2行为表头
        except FileNotFoundError:
            raise ValueError("Template file not found.")
        except InvalidFileException:
            raise ValueError("Invalid template file format.")
    elif headers:
        # 设置表头
        for col_idx, header in enumerate(headers, start=1):
            header_name = header if isinstance(header, str) else header.get("name", "")
            header_id = header if isinstance(header, str) else header.get("id", "")
            cell = sheet.cell(row=1, column=col_idx, value=str(header_name))
            if header_style and header_id in header_style:
                # 表头格式设置为粗体，还需要设置颜色
                cell.fill = PatternFill("solid", fgColor=header_style[header_id])
    else:
        raise ValueError("Either template or headers must be provided.")

    if match_header and not headers:
        raise ValueError("Headers must be provided when match_header is True.")

    # 数据写入单元格
    for row_idx, data_dict in enumerate(data_dict__list, start=1):
        sheet_row_idx = row_idx + first_data_row - 1
        if match_header:
            for col_idx, header in enumerate(headers, start=1):
                header_key = header if isinstance(header, str) else header.get("id", "")
                cell = sheet.cell(row=sheet_row_idx, column=col_idx, value="")
                if header_key in data_dict:
                    cell.value = str(data_dict[header_key])
        else:
            for col_idx, value in enumerate(data_dict.values(), start=1):
                sheet.cell(row=sheet_row_idx, column=col_idx, value=str(value))

    # 自适应行高和列宽
    adapt_sheet_weight_height(sheet=sheet, first_header_row=1 if not template else 2)

    return wb
